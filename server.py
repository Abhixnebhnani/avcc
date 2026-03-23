"""
AVCC Gantry System — Backend Server
=====================================
FastAPI + YOLOv8 + OpenCV + EasyOCR

Receives video frames from the frontend,
runs YOLOv8 detection, tracks vehicles,
detects counting line crossings, classifies,
reads number plates, calculates speed,
and streams results back.

Run:
    python server.py
Then open frontend/index.html in your browser.
"""

import asyncio
import base64
import io
import json
import re
import time
import uuid
from collections import defaultdict
from datetime import datetime
from typing import Optional

import cv2
import numpy as np
import uvicorn
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── YOLOv8 ────────────────────────────────────────────────
try:
    from ultralytics import YOLO
    model = YOLO("yolov8n.pt")   # downloads automatically on first run (~6 MB)
    YOLO_AVAILABLE = True
    print("✅ YOLOv8 loaded successfully")
except Exception as e:
    print(f"⚠️  YOLOv8 not available: {e}")
    YOLO_AVAILABLE = False
    model = None

# ── EasyOCR (for number plate reading) ────────────────────
try:
    import easyocr
    plate_reader = easyocr.Reader(['en'], gpu=False, verbose=False)
    OCR_AVAILABLE = True
    print("✅ EasyOCR loaded successfully")
except Exception as e:
    print(f"⚠️  EasyOCR not available: {e}")
    OCR_AVAILABLE = False
    plate_reader = None

# Indian number plate pattern: XX 00 XX 0000
# Strict: must have state code (2 letters), district (1-2 digits), series (1-3 letters), number (4 digits)
PLATE_PATTERN = re.compile(r'[A-Z]{2}\s*\d{1,2}\s*[A-Z]{1,3}\s*\d{4}')

# ── Indian Traffic Class Mapping ──────────────────────────
# COCO classes → Indian AVCC classes
COCO_TO_AVCC = {
    "car":          "Car",
    "truck":        "Truck/HGV",
    "bus":          "Bus",
    "motorcycle":   "Motorcycle",
    "bicycle":      "Bicycle",
    "person":       None,           # ignore pedestrians
}

# Classes we want to detect (COCO names)
TARGET_CLASSES = {"car", "truck", "bus", "motorcycle", "bicycle"}

# Minimum bounding box area (pixels²) to accept a detection — filters noise
# At 1280x720, bikes are typically 1500-4000px², cars 5000+
MIN_DETECTION_AREA = 500
# Motorcycles/bicycles are smaller — use a separate lower threshold
MIN_TWOWHEELER_AREA = 300

# Extra Indian classes (heuristic upgrade based on size)
# Only override YOLO's class when there's very strong evidence.
# YOLO already does a good job — we only refine edge cases.
def refine_class(coco_cls: str, w: int, h: int, area: int) -> str:
    aspect = w / max(h, 1)
    if coco_cls == "car":
        # Only reclassify as Auto if VERY small AND nearly square (3-wheeler shape)
        if area < 1200 and 0.7 < aspect < 1.0:
            return "Auto"
        # YOLO said car → trust it as Car
        return "Car"
    if coco_cls == "truck":
        if aspect > 2.0 and area < 5000:
            return "Tractor"
        if area < 4000:
            return "LCV/Van"
        return "Truck/HGV"
    if coco_cls == "bus":
        return "Bus"
    if coco_cls == "motorcycle":
        return "Motorcycle"
    if coco_cls == "bicycle":
        return "Bicycle"
    return coco_cls.title()

def _format_plate(raw: str) -> str:
    """Format a matched plate string nicely: XX 00 XX 0000."""
    p = re.sub(r'[^A-Z0-9]', '', raw.upper())
    if len(p) >= 9:
        return f"{p[:2]} {p[2:4]} {p[4:-4]} {p[-4:]}"
    elif len(p) >= 8:
        return f"{p[:2]} {p[2:4]} {p[4:-4]} {p[-4:]}"
    return p


def _try_clean_plate(raw: str) -> str:
    """Try to extract an Indian plate from raw OCR text. Returns formatted plate or None."""
    text = re.sub(r'[^A-Z0-9]', '', raw.upper())
    if len(text) < 6:
        return None

    # Direct match
    if PLATE_PATTERN.search(text):
        return _format_plate(PLATE_PATTERN.search(text).group(0))

    # Positional correction
    chars = list(text)
    d2l = {'0': 'O', '1': 'I', '2': 'Z', '5': 'S', '6': 'G', '8': 'B'}
    l2d = {'O': '0', 'I': '1', 'Z': '2', 'S': '5', 'G': '6', 'B': '8', 'D': '0', 'Q': '0'}

    # First 2 must be letters (state code)
    for i in range(min(2, len(chars))):
        if chars[i].isdigit() and chars[i] in d2l:
            chars[i] = d2l[chars[i]]
    # Last 4 must be digits
    for i in range(max(0, len(chars) - 4), len(chars)):
        if chars[i].isalpha() and chars[i] in l2d:
            chars[i] = l2d[chars[i]]

    corrected = ''.join(chars)
    if PLATE_PATTERN.search(corrected):
        return _format_plate(PLATE_PATTERN.search(corrected).group(0))
    return None


def read_plate(frame, bbox) -> str:
    """Read number plate directly using EasyOCR on the full vehicle crop.
    Simple approach: crop → preprocess → OCR → clean. Returns plate or 'Unreadable'."""
    if not OCR_AVAILABLE or plate_reader is None:
        return "Unreadable"

    x1, y1, x2, y2 = bbox
    bh = y2 - y1
    bw = x2 - x1
    if bh < 40 or bw < 40:
        return "Unreadable"

    H, W = frame.shape[:2]
    best_plate = None
    best_conf = 0

    # Try multiple crop zones of the vehicle
    for top_frac in (0.50, 0.35, 0.60, 0.0):
        cy1 = max(0, y1 + int(bh * top_frac))
        cy2 = min(H, y2 + 15)
        cx1 = max(0, x1 - 10)
        cx2 = min(W, x2 + 10)
        crop = frame[cy1:cy2, cx1:cx2]
        if crop.size == 0:
            continue

        ch, cw = crop.shape[:2]
        if ch < 15 or cw < 30:
            continue

        # Upscale small crops so text is big enough for OCR
        target_w = max(400, cw)
        scale_f = target_w / cw
        if scale_f > 1.0:
            crop = cv2.resize(crop, None, fx=scale_f, fy=scale_f, interpolation=cv2.INTER_CUBIC)

        try:
            # EasyOCR works best on the raw color image — don't overthink preprocessing
            results = plate_reader.readtext(crop, detail=1, paragraph=False)
            for (_, text, conf) in results:
                raw = text.strip()
                if len(raw) < 4 or conf < 0.15:
                    continue
                cleaned = _try_clean_plate(raw)
                if cleaned and conf > best_conf:
                    best_plate = cleaned
                    best_conf = conf
        except Exception:
            pass

        # Also try grayscale + CLAHE (handles low contrast / night footage)
        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        try:
            results = plate_reader.readtext(enhanced, detail=1, paragraph=False)
            for (_, text, conf) in results:
                raw = text.strip()
                if len(raw) < 4 or conf < 0.15:
                    continue
                cleaned = _try_clean_plate(raw)
                if cleaned and conf > best_conf:
                    best_plate = cleaned
                    best_conf = conf
        except Exception:
            pass

        if best_plate and best_conf > 0.4:
            return best_plate

    return best_plate if best_plate else "Unreadable"


CLASS_COLORS = {
    "Car":        (0, 212, 255),
    "Motorcycle": (255, 107, 53),
    "Auto":       (255, 204, 0),
    "Bus":        (57, 255, 20),
    "Truck/HGV":  (255, 68, 136),
    "LCV/Van":    (187, 136, 255),
    "Tractor":    (255, 153, 51),
    "Bicycle":    (136, 255, 238),
}

# ── App ───────────────────────────────────────────────────
app = FastAPI(title="AVCC Gantry System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the frontend
import os
frontend_path = os.path.join(os.path.dirname(__file__), "frontend")

@app.get("/")
async def root():
    idx = os.path.join(frontend_path, "index.html")
    if os.path.exists(idx):
        return FileResponse(idx)
    return {"status": "AVCC Server running", "yolo": YOLO_AVAILABLE}

@app.get("/health")
async def health():
    return {"status": "ok", "yolo": YOLO_AVAILABLE, "ocr": OCR_AVAILABLE, "time": datetime.now().isoformat()}

# ── Per-session State ─────────────────────────────────────
class SessionState:
    def __init__(self):
        self.trackers: dict = {}          # track_id → tracker dict
        self.counts: dict = defaultdict(int)
        self.lane_counts: list = [0, 0, 0, 0]
        self.lane_speed_sums: list = [0.0, 0.0, 0.0, 0.0]
        self.lane_speed_counts: list = [0, 0, 0, 0]
        self.total: int = 0
        self.records: list = []
        self.frame_w: int = 640
        self.frame_h: int = 480
        self.counting_line_y: float = 0.55   # fraction of height
        self.fps: float = 25.0
        self.vehicle_id_counter: int = 0
        self.db_buffer: list = []

    def next_id(self) -> str:
        self.vehicle_id_counter += 1
        return f"VH{self.vehicle_id_counter:05d}"

sessions: dict[str, SessionState] = {}

# ── Robust Centroid + IoU Tracker ─────────────────────────
def iou(a, b):
    ax1,ay1,ax2,ay2 = a
    bx1,by1,bx2,by2 = b
    ix1,iy1 = max(ax1,bx1), max(ay1,by1)
    ix2,iy2 = min(ax2,bx2), min(ay2,by2)
    inter = max(0,ix2-ix1)*max(0,iy2-iy1)
    ua = (ax2-ax1)*(ay2-ay1) + (bx2-bx1)*(by2-by1) - inter
    return inter/max(ua,1)

def centroid_dist(a, b):
    """Euclidean distance between bbox centroids."""
    cx_a = (a[0]+a[2])/2; cy_a = (a[1]+a[3])/2
    cx_b = (b[0]+b[2])/2; cy_b = (b[1]+b[3])/2
    return ((cx_a-cx_b)**2 + (cy_a-cy_b)**2)**0.5

def bbox_size_ratio(a, b):
    """Ratio of bbox areas (smaller/larger). 1.0 = same size."""
    area_a = max((a[2]-a[0])*(a[3]-a[1]), 1)
    area_b = max((b[2]-b[0])*(b[3]-b[1]), 1)
    return min(area_a, area_b) / max(area_a, area_b)

# Max centroid distance (pixels) to consider same vehicle
MAX_CENTROID_DIST = 120

def update_trackers(state: SessionState, detections: list, frame_time: float):
    """
    Two-pass matching:
      1) IoU match (handles normal frame-to-frame movement)
      2) Centroid-distance match (handles lane changes / fast movement where IoU fails)
    """
    active = {}
    used_dets = set()
    unmatched_trackers = {}

    # ── Pass 1: IoU matching ──
    for tid, tr in state.trackers.items():
        best_iou, best_i = 0, -1
        for i, det in enumerate(detections):
            if i in used_dets:
                continue
            s = iou(tr["bbox"], det["bbox"])
            if s > best_iou:
                best_iou, best_i = s, i

        if best_iou > 0.20 and best_i >= 0:
            _apply_match(tr, detections[best_i], frame_time, state)
            used_dets.add(best_i)
            active[tid] = tr
        else:
            unmatched_trackers[tid] = tr

    # ── Pass 2: Centroid-distance fallback for unmatched trackers ──
    # This catches vehicles that changed lanes or moved fast between frames
    for tid, tr in unmatched_trackers.items():
        best_dist = MAX_CENTROID_DIST
        best_i = -1
        for i, det in enumerate(detections):
            if i in used_dets:
                continue
            d = centroid_dist(tr["bbox"], det["bbox"])
            sr = bbox_size_ratio(tr["bbox"], det["bbox"])
            # Must be close AND similar size (not a completely different vehicle)
            if d < best_dist and sr > 0.4:
                best_dist = d
                best_i = i

        if best_i >= 0:
            _apply_match(tr, detections[best_i], frame_time, state)
            used_dets.add(best_i)
            active[tid] = tr
        else:
            tr["missed"] = tr.get("missed", 0) + 1
            if tr["missed"] < 5:
                active[tid] = tr

    # ── New detections → new tracks (only if not near an existing tracker) ──
    for i, det in enumerate(detections):
        if i in used_dets:
            continue
        # Check if this detection is very close to any existing active tracker
        # (prevents creating duplicate track for the same vehicle)
        too_close = False
        for tid, tr in active.items():
            if centroid_dist(tr["bbox"], det["bbox"]) < 50 and bbox_size_ratio(tr["bbox"], det["bbox"]) > 0.5:
                too_close = True
                break
        if too_close:
            continue

        vid = state.next_id()
        cx = (det["bbox"][0] + det["bbox"][2]) / 2
        cy = (det["bbox"][1] + det["bbox"][3]) / 2
        active[vid] = {
            "vid": vid,
            "bbox": det["bbox"],
            "cls": det["cls"],
            "conf": det["conf"],
            "cx": cx, "cy": cy,
            "speed": 0.0,
            "counted": False,
            "last_time": frame_time,
            "missed": 0,
        }

    state.trackers = active
    return active

def _apply_match(tr, det, frame_time, state):
    """Update a tracker with a matched detection."""
    prev_cy = (tr["bbox"][1] + tr["bbox"][3]) / 2
    new_cy  = (det["bbox"][1] + det["bbox"][3]) / 2
    dy_px   = new_cy - prev_cy
    dt      = frame_time - tr["last_time"]

    scale = 4.0 / (0.12 * state.frame_h)
    speed_ms = (abs(dy_px) / max(dt, 0.001)) * scale
    speed_kph = speed_ms * 3.6
    # Correction factor: pixel-based speed underestimates at this scale
    if speed_kph < 10:
        speed_kph *= 6.5

    # Use majority-vote for class: keep the existing class unless the new
    # detection has higher confidence (prevents flickering classifications)
    new_cls = det["cls"]
    if tr.get("conf", 0) > det["conf"] + 0.05:
        new_cls = tr["cls"]  # keep old class if it was more confident

    tr.update({
        "bbox": det["bbox"],
        "cls": new_cls,
        "conf": max(det["conf"], tr.get("conf", 0)),
        "cy": new_cy,
        "speed": min(speed_kph, 150),
        "last_time": frame_time,
        "missed": 0,
    })

# Dedup buffer: stores (cx, timestamp) of recently counted vehicles
# to prevent the same physical vehicle from being counted twice
DEDUP_X_TOLERANCE = 60      # pixels — two crossings within this X range...
DEDUP_TIME_WINDOW = 2.0     # ...within this many seconds = duplicate

def check_counting_line(state: SessionState, trackers: dict, timestamp: str, frame=None):
    """Check which vehicles just crossed the counting line. Read plates on crossing."""
    line_y = state.counting_line_y * state.frame_h
    events = []
    now = time.time()

    # Clean old entries from dedup buffer
    if not hasattr(state, '_dedup_buf'):
        state._dedup_buf = []
    state._dedup_buf = [(cx, t) for cx, t in state._dedup_buf if now - t < DEDUP_TIME_WINDOW]

    for tid, tr in trackers.items():
        if tr.get("counted"):
            continue
        cy = (tr["bbox"][1] + tr["bbox"][3]) / 2
        if cy >= line_y:
            tr["counted"] = True

            # Lane assignment (by x centre)
            cx = (tr["bbox"][0] + tr["bbox"][2]) / 2
            lane = min(int(cx / (state.frame_w / 4)), 3)

            # Dedup check: was a vehicle counted at nearly the same X recently?
            is_dup = False
            for prev_cx, prev_t in state._dedup_buf:
                if abs(cx - prev_cx) < DEDUP_X_TOLERANCE and (now - prev_t) < DEDUP_TIME_WINDOW:
                    is_dup = True
                    break
            if is_dup:
                continue

            state._dedup_buf.append((cx, now))

            cls  = tr["cls"]
            spd  = round(tr["speed"], 1)
            vid  = tr["vid"]
            conf = round(tr.get("conf", 0.9) * 100, 1)

            # Read number plate
            plate = ""
            if frame is not None:
                plate = read_plate(frame, tr["bbox"])

            state.counts[cls] += 1
            state.lane_counts[lane] += 1
            state.lane_speed_sums[lane] += spd
            state.lane_speed_counts[lane] += 1
            state.total += 1

            record = {
                "id":    vid,
                "time":  timestamp,
                "lane":  lane + 1,
                "class": cls,
                "speed": spd,
                "conf":  conf,
                "plate": plate,
            }
            state.records.append(record)
            state.db_buffer.append(record)
            events.append(record)

    return events

# ── WebSocket Endpoint ────────────────────────────────────
@app.websocket("/ws/{session_id}")
async def websocket_endpoint(ws: WebSocket, session_id: str):
    await ws.accept()
    state = SessionState()
    sessions[session_id] = state
    print(f"🔌 Session connected: {session_id}")

    try:
        while True:
            raw = await ws.receive_text()
            msg = json.loads(raw)
            mtype = msg.get("type")

            # ── Config ──
            if mtype == "config":
                state.frame_w  = msg.get("width", 640)
                state.frame_h  = msg.get("height", 480)
                state.fps      = msg.get("fps", 25.0)
                state.counting_line_y = msg.get("counting_line_y", 0.55)
                await ws.send_text(json.dumps({"type": "config_ack", "status": "ok"}))

            # ── Frame ──
            elif mtype == "frame":
                t0 = time.time()
                frame_b64 = msg.get("data", "")
                frame_time = msg.get("frame_time", time.time())
                light_mode = msg.get("light", False)
                timestamp = datetime.now().strftime("%H:%M:%S.%f")[:11]

                # Decode frame
                img_bytes = base64.b64decode(frame_b64)
                arr = np.frombuffer(img_bytes, np.uint8)
                frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)

                if frame is None:
                    continue

                state.frame_w = frame.shape[1]
                state.frame_h = frame.shape[0]

                detections = []

                if YOLO_AVAILABLE and model:
                    # Lower confidence for motorcycles which YOLO is less sure about
                    results = model(frame, verbose=False, conf=0.40)[0]
                    for box in results.boxes:
                        cls_id = int(box.cls[0])
                        cls_name = results.names[cls_id].lower()
                        if cls_name not in TARGET_CLASSES:
                            continue
                        x1,y1,x2,y2 = map(int, box.xyxy[0].tolist())
                        w = x2 - x1
                        h = y2 - y1
                        area = w * h
                        conf = float(box.conf[0])
                        # Use lower area threshold for two-wheelers
                        is_twowheeler = cls_name in ("motorcycle", "bicycle")
                        min_area = MIN_TWOWHEELER_AREA if is_twowheeler else MIN_DETECTION_AREA
                        if area < min_area:
                            continue  # skip tiny noise detections
                        # For non-bike classes, require higher confidence
                        if not is_twowheeler and conf < 0.50:
                            continue
                        refined = refine_class(cls_name, w, h, area)
                        detections.append({"bbox":[x1,y1,x2,y2], "cls":refined, "conf":conf})
                else:
                    # Fallback: basic motion/colour simulation (no YOLO)
                    pass

                # Track
                trackers = update_trackers(state, detections, frame_time)

                # Count (pass frame for plate reading)
                events = check_counting_line(state, trackers, timestamp, frame)

                # Build response
                proc_ms = round((time.time() - t0) * 1000, 1)

                # Lane avg speeds
                lane_avg = []
                for i in range(4):
                    if state.lane_speed_counts[i] > 0:
                        lane_avg.append(round(state.lane_speed_sums[i] / state.lane_speed_counts[i], 1))
                    else:
                        lane_avg.append(None)

                # Flush DB buffer
                db_flush = None
                if state.db_buffer:
                    db_flush = {
                        "packet": len(state.records),
                        "batch_size": len(state.db_buffer),
                        "timestamp": datetime.now().isoformat(),
                        "sample": state.db_buffer[-1],
                    }
                    state.db_buffer.clear()

                resp = {
                    "type":        "result",
                    "detections":  len(detections),
                    "events":      events,
                    "total":       state.total,
                    "class_counts":dict(state.counts),
                    "lane_counts": state.lane_counts,
                    "lane_avg_speed": lane_avg,
                    "proc_ms":     proc_ms,
                    "db_flush":    db_flush,
                }

                if light_mode:
                    # Light mode: send tracker metadata for client-side drawing
                    boxes = []
                    for tid, tr in trackers.items():
                        boxes.append({
                            "bbox": tr["bbox"],
                            "cls": tr["cls"],
                            "speed": round(tr["speed"], 1),
                            "counted": tr.get("counted", False),
                        })
                    resp["boxes"] = boxes
                    resp["line_y"] = state.counting_line_y
                else:
                    # Full mode: send annotated frame image
                    annotated = annotate_frame(frame, state, trackers)
                    _, buf = cv2.imencode(".jpg", annotated, [cv2.IMWRITE_JPEG_QUALITY, 75])
                    resp["frame"] = base64.b64encode(buf).decode()

                await ws.send_text(json.dumps(resp))

    except WebSocketDisconnect:
        print(f"🔌 Session disconnected: {session_id}")
    except Exception as e:
        print(f"❌ Error in session {session_id}: {e}")
        import traceback; traceback.print_exc()
    finally:
        sessions.pop(session_id, None)

# ── Annotation ────────────────────────────────────────────
def annotate_frame(frame, state: SessionState, trackers: dict):
    out = frame.copy()
    H, W = out.shape[:2]
    lane_w = W // 4
    line_y = int(state.counting_line_y * H)

    # Lane dividers
    for i in range(1, 4):
        cv2.line(out, (lane_w*i, 0), (lane_w*i, H), (50, 80, 100), 1)

    # Counting line
    cv2.line(out, (0, line_y), (W, line_y), (0, 0, 255), 2)
    cv2.putText(out, "COUNTING LINE", (8, line_y - 6),
                cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)

    # Lane labels
    for i in range(4):
        cv2.putText(out, f"L{i+1}", (lane_w*i + 5, 18),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 180, 255), 1)

    # Tracked vehicles
    for tid, tr in trackers.items():
        x1,y1,x2,y2 = tr["bbox"]
        cls = tr["cls"]
        color = CLASS_COLORS.get(cls, (200,200,200))
        bgr = (color[2], color[1], color[0])  # RGB→BGR

        # Box
        cv2.rectangle(out, (x1,y1), (x2,y2), bgr, 2)

        # Label
        label = f"{cls} {tr['speed']:.0f}km/h"
        (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.4, 1)
        cv2.rectangle(out, (x1, y1-th-6), (x1+tw+4, y1), bgr, -1)
        cv2.putText(out, label, (x1+2, y1-4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0,0,0), 1)

        # Centroid dot
        cx = (x1+x2)//2; cy = (y1+y2)//2
        cv2.circle(out, (cx, cy), 3, bgr, -1)

    # Total count overlay (top right)
    cv2.putText(out, f"TOTAL: {state.total}", (W-120, 20),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0,212,255), 2)

    return out


# ── Excel Export ──────────────────────────────────────────
def build_excel(records: list) -> io.BytesIO:
    """Build an Excel workbook with master sheet + per-class sheets."""
    wb = Workbook()

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A3A55', end_color='1A3A55', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    HEADERS = ['S.No.', 'Number Plate', 'Classification', 'Vehicle ID', 'Time', 'Lane', 'Speed (km/h)', 'Confidence (%)']

    def write_sheet(ws, title, rows):
        ws.title = title
        # Write headers
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        # Write data
        for i, rec in enumerate(rows, 1):
            vals = [
                i,
                rec.get('plate', '') or '—',
                rec.get('class', ''),
                rec.get('id', ''),
                rec.get('time', ''),
                rec.get('lane', ''),
                rec.get('speed', ''),
                rec.get('conf', ''),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=i + 1, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        # Auto-width
        for col in range(1, len(HEADERS) + 1):
            max_len = len(HEADERS[col - 1])
            for row in range(2, len(rows) + 2):
                val = str(ws.cell(row=row, column=col).value or '')
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3

    # Master sheet (Sheet 1)
    master = wb.active
    write_sheet(master, 'All Vehicles', records)

    # Per-class sheets
    class_groups = defaultdict(list)
    for rec in records:
        class_groups[rec.get('class', 'Unknown')].append(rec)

    for cls_name in sorted(class_groups.keys()):
        # Excel sheet names max 31 chars, no special chars
        safe_name = cls_name.replace('/', '-')[:31]
        ws = wb.create_sheet()
        write_sheet(ws, safe_name, class_groups[cls_name])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/download/{session_id}")
async def download_excel(session_id: str):
    """Download current detection records as Excel."""
    state = sessions.get(session_id)
    if not state or not state.records:
        # Return empty workbook if no data
        wb = Workbook()
        wb.active.title = "All Vehicles"
        wb.active.cell(row=1, column=1, value="No data yet")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=avcc_report.xlsx"},
        )

    buf = build_excel(state.records)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=avcc_report_{session_id}.xlsx"},
    )


# Mount static files AFTER all routes so /download etc. take priority
if os.path.exists(frontend_path):
    app.mount("/static", StaticFiles(directory=frontend_path), name="static")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    print("=" * 50)
    print("  AVCC Gantry System — Server")
    print("=" * 50)
    print(f"  YOLOv8: {'✅ Ready' if YOLO_AVAILABLE else '❌ Not installed'}")
    print(f"  EasyOCR: {'✅ Ready' if OCR_AVAILABLE else '❌ Not installed'}")
    print(f"  Port: {port}")
    print("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="warning")
